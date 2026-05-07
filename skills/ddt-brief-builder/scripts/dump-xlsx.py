#!/usr/bin/env python3
"""
ddt-brief-builder · xlsx 全文 dump

Read 工具不支持 xlsx，本脚本用 openpyxl 把 xlsx 全文 dump 为
"行号 | A | B | C | ..." 的文本格式，让 LLM 直接读。

设计原则（v0.9.4 丝滑 KPI）：
- openpyxl 缺失自动 pip install（不让用户中断）
- 单 sheet / 多 sheet 都支持
- 默认全文（不截断）让 LLM 完整提取
- 失败信息直接包含修复命令

用法：
    python3 dump-xlsx.py <path-to.xlsx>
    python3 dump-xlsx.py <path-to.xlsx> --sheet Sheet1
    python3 dump-xlsx.py <path-to.xlsx> --max-rows 50    # 大文件截断

退出码：
    0 = dump 成功
    1 = 文件不存在 / 不是 xlsx
    2 = openpyxl 安装失败
    3 = xlsx 解析失败（损坏或格式不对）
"""
from __future__ import annotations
import argparse
import subprocess
import sys
from pathlib import Path


def ensure_openpyxl() -> None:
    """openpyxl 缺失自动 pip install，让 skill 调用方不必预装。"""
    try:
        import openpyxl  # noqa: F401
        return
    except ImportError:
        pass
    print("⚙️  openpyxl 未安装，自动 pip install...", file=sys.stderr)
    cmds = [
        [sys.executable, "-m", "pip", "install", "--quiet", "openpyxl"],
        [sys.executable, "-m", "pip", "install", "--user", "--quiet", "openpyxl"],
    ]
    for cmd in cmds:
        try:
            subprocess.run(cmd, check=True, capture_output=True)
            import openpyxl  # noqa: F401
            print("✅ openpyxl 安装成功", file=sys.stderr)
            return
        except (subprocess.CalledProcessError, ImportError):
            continue
    print("❌ openpyxl 安装失败。手动运行：", file=sys.stderr)
    print("   pip3 install openpyxl    或    python3 -m pip install --user openpyxl", file=sys.stderr)
    sys.exit(2)


def dump_xlsx(path: Path, sheet_name: str | None = None, max_rows: int | None = None) -> None:
    import openpyxl

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"❌ xlsx 解析失败：{e}", file=sys.stderr)
        print(f"   可能原因：文件损坏 / 受密码保护 / 不是真正的 xlsx 格式", file=sys.stderr)
        sys.exit(3)

    sheets = [sheet_name] if sheet_name else wb.sheetnames
    for name in sheets:
        if name not in wb.sheetnames:
            print(f"❌ sheet '{name}' 不存在。可用 sheet：{wb.sheetnames}", file=sys.stderr)
            sys.exit(1)
        ws = wb[name]
        print(f"--- Sheet: {name} ({ws.max_row} 行 x {ws.max_column} 列) ---")
        truncated = False
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if max_rows and i > max_rows:
                truncated = True
                break
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                print(f"{i:3d} | " + " | ".join(cells))
        if truncated:
            print(f"... (省略，--max-rows {max_rows} 已截断；不传 --max-rows 看完整)")
        print()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="ddt-brief-builder xlsx 全文 dump（自动 pip install openpyxl）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("path", help="xlsx 文件路径（绝对或相对）")
    parser.add_argument("--sheet", help="指定 sheet（默认 dump 所有）")
    parser.add_argument("--max-rows", type=int, help="单 sheet 截断行数（默认全文）")
    args = parser.parse_args()

    path = Path(args.path).expanduser()
    if not path.exists():
        print(f"❌ 文件不存在：{path}", file=sys.stderr)
        return 1
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        print(f"❌ 仅支持 .xlsx/.xlsm，实际：{path.suffix}", file=sys.stderr)
        print(f"   .xls 老格式请先在 Excel 里另存为 .xlsx", file=sys.stderr)
        return 1

    ensure_openpyxl()
    dump_xlsx(path, sheet_name=args.sheet, max_rows=args.max_rows)
    return 0


if __name__ == "__main__":
    sys.exit(main())
