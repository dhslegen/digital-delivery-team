#!/usr/bin/env python3
"""
ddt-baseline-sync · 人员需求计划表解析器

把 xlsx 人员表解析为结构化 JSON，按角色 → DDT phase 映射，输出 CSV 行所需的所有字段。

用法：
    python3 parse-staffing.py <path-to.xlsx>
    python3 parse-staffing.py <path-to.xlsx> --sheet Sheet1

输出（stdout 单 JSON）：
    {
      "project_name": "无人物流车运营系统",
      "team_size": 7,
      "total_person_months": 5.0,
      "total_hours": 880,
      "complexity": "复杂",
      "time_window": "2026-01-04 ~ 2026-03-15",
      "roles": [
        {"name": "韩金群", "role": "项目经理/架构工程师", "person_months": 1.4, ...}
      ],
      "phase_hours": {
        "requirements_hours": 140,
        "wbs_hours": 0,
        "design_hours": 246,
        "ui_design_hours": 53,
        "frontend_hours": 211,
        "backend_hours": 140,
        "test_hours": 62,
        "review_hours": 26,
        "docs_hours": 0
      }
    }

退出码：0 成功；1 参数错误；2 openpyxl 安装失败；3 解析失败
"""
from __future__ import annotations
import argparse
import json
import re
import subprocess
import sys
from pathlib import Path
from typing import Optional


def ensure_openpyxl() -> None:
    try:
        import openpyxl  # noqa: F401
        return
    except ImportError:
        pass
    print("⚙️  openpyxl 未安装，自动 pip install...", file=sys.stderr)
    cmds = [
        [sys.executable, "-m", "pip", "install", "--quiet", "openpyxl"],
        [sys.executable, "-m", "pip", "install", "--user", "--quiet", "openpyxl"],
    ]
    for cmd in cmds:
        try:
            subprocess.run(cmd, check=True, capture_output=True)
            import openpyxl  # noqa: F401
            return
        except (subprocess.CalledProcessError, ImportError):
            continue
    print("❌ openpyxl 安装失败。手动: pip3 install openpyxl", file=sys.stderr)
    sys.exit(2)


# 角色 → phase 映射（与 SKILL.md 对齐）
def role_to_phase_split(role: str) -> dict[str, float]:
    """返回该角色在各 phase 的人月权重（汇总所有人月）"""
    role = role.strip()
    if "项目经理" in role and "架构" in role:
        return {"architecture": 0.6, "requirements": 0.4}
    if "项目经理" in role:
        return {"requirements": 1.0}
    if "架构" in role:
        return {"architecture": 1.0}
    if "产品" in role or "PM" in role.upper():
        return {"requirements": 0.5, "wbs": 0.5}  # prd + wbs 各半
    if "UI" in role.upper() or "设计" in role:
        return {"ui_design": 1.0}
    if "前端" in role or "fe" in role.lower():
        return {"frontend": 1.0}
    if "后端" in role or "be" in role.lower():
        return {"backend": 1.0}
    if "测试" in role or "QA" in role.upper() or "test" in role.lower():
        return {"test": 0.7, "review": 0.3}
    if "运维" in role or "DevOps" in role or "SRE" in role.upper():
        return {"docs": 1.0}
    return {"docs": 1.0}  # 兜底归 docs


def parse_date(value):
    if value is None: return None
    s = str(value).split(" ")[0]  # "2026-01-04 00:00:00" → "2026-01-04"
    return s


def parse_xlsx(path: Path, sheet_name: Optional[str] = None) -> dict:
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"❌ xlsx 解析失败: {e}", file=sys.stderr)
        sys.exit(3)

    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    # 提取项目名（第 1 行第 2 列通常是值）
    project_name = ""
    for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
        for i, c in enumerate(row):
            if c and "项目名称" in str(c) and i + 1 < len(row) and row[i + 1]:
                project_name = str(row[i + 1]).strip()
                break
        if project_name: break

    # 找到表头行（含"姓名" + "角色" + "人月"）
    header_row = None
    header_indices = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [str(c) if c else "" for c in row]
        if any("姓名" in c for c in cells) and any("角色" in c for c in cells) and any("人月" in c for c in cells):
            header_row = i
            for j, c in enumerate(cells):
                if "姓名" in c: header_indices["name"] = j
                elif "部门" in c: header_indices["dept"] = j
                elif "角色" in c: header_indices["role"] = j
                elif "级别" in c: header_indices["level"] = j
                elif "进入时间" in c or "进入" in c: header_indices["start"] = j
                elif "离开时间" in c or "离开" in c: header_indices["end"] = j
                elif "人月" in c: header_indices["pm"] = j
                elif "工作内容" in c: header_indices["task"] = j
            break

    if not header_row:
        print("❌ 未找到含'姓名 + 角色 + 人月'的表头行", file=sys.stderr)
        sys.exit(3)

    # 解析数据行
    roles = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        cells = list(row)
        if not cells or all(c is None for c in cells): continue
        get = lambda key: cells[header_indices[key]] if key in header_indices and header_indices[key] < len(cells) else None
        name = get("name")
        role = get("role")
        pm = get("pm")
        if not name or not role or pm is None: continue
        try:
            pm_num = float(pm)
        except (TypeError, ValueError):
            continue
        roles.append({
            "name": str(name).strip(),
            "department": str(get("dept") or "").strip(),
            "role": str(role).strip(),
            "level": str(get("level") or "").strip(),
            "start_date": parse_date(get("start")),
            "end_date": parse_date(get("end")),
            "person_months": pm_num,
            "task": str(get("task") or "").strip(),
        })

    if not roles:
        print("❌ 未解析出任何角色行", file=sys.stderr)
        sys.exit(3)

    # 累加 phase 工时
    phase_hours = {
        "requirements_hours": 0, "wbs_hours": 0,
        "architecture_hours": 0, "ui_design_hours": 0,
        "frontend_hours": 0, "backend_hours": 0,
        "test_hours": 0, "review_hours": 0, "docs_hours": 0,
    }
    for r in roles:
        split = role_to_phase_split(r["role"])
        for phase_key, weight in split.items():
            full_key = f"{phase_key}_hours"
            if full_key in phase_hours:
                phase_hours[full_key] += r["person_months"] * 176 * weight

    phase_hours = {k: round(v) for k, v in phase_hours.items()}
    # design_hours = architecture + ui_design 合并（baseline schema 单列）
    design_total = phase_hours["architecture_hours"] + phase_hours["ui_design_hours"]

    total_pm = round(sum(r["person_months"] for r in roles), 2)
    total_hours = round(total_pm * 176)

    if total_hours <= 60: complexity = "简单"
    elif total_hours <= 200: complexity = "中等"
    else: complexity = "复杂"

    starts = [r["start_date"] for r in roles if r["start_date"]]
    ends = [r["end_date"] for r in roles if r["end_date"]]
    time_window = ""
    if starts and ends:
        time_window = f"{min(starts)} ~ {max(ends)}"

    return {
        "project_name": project_name,
        "team_size": len(roles),
        "total_person_months": total_pm,
        "total_hours": total_hours,
        "complexity": complexity,
        "time_window": time_window,
        "roles": roles,
        "phase_hours": {
            "prd_hours": phase_hours["requirements_hours"],
            "wbs_hours": phase_hours["wbs_hours"],
            "design_hours": design_total,
            "frontend_hours": phase_hours["frontend_hours"],
            "backend_hours": phase_hours["backend_hours"],
            "test_hours": phase_hours["test_hours"],
            "review_hours": phase_hours["review_hours"],
            "docs_hours": phase_hours["docs_hours"],
        },
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="人员需求表 → DDT baseline JSON")
    parser.add_argument("path", help="xlsx 路径")
    parser.add_argument("--sheet", help="指定 sheet（默认第一个）")
    args = parser.parse_args()

    path = Path(args.path).expanduser()
    if not path.exists():
        print(f"❌ 文件不存在: {path}", file=sys.stderr)
        return 1
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        print(f"❌ 仅支持 xlsx/xlsm: {path.suffix}", file=sys.stderr)
        return 1

    ensure_openpyxl()
    result = parse_xlsx(path, args.sheet)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
